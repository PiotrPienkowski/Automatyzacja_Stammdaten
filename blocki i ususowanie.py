
import time
import os
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook

template = r'C:\Users\02703821\Elanco\CH - Bestellung Monitoring\Templatka do pythona GTS\(de and ch sold-to block unblock).xlsm'
new_file = r"C:\Users\02703821\OneDrive - Elanco\Desktop\robocze"
snow = "https://thespot.elanco.com/esc?id=sc_cat_item&sys_id=9d661f191b03d1105ca7eca3604bcb3a&sysparm_category=a20cb8eedb7c60905513c3af299619d0"
snow_ch = "https://thespot.elanco.com/esc?id=sc_cat_item&table=sc_cat_item&sys_id=666af21697260e907487fd7ef053afd3&recordUrl=com.glideapp.servicecatalog_cat_item_view.do%3Fv%3D1&sysparm_id=666af21697260e907487fd7ef053afd3"

os.system("taskkill /F /IM excel.exe >nul 2>nul")

class snow_ticket:

    def __init__(self):

        self.pw = sync_playwright().start()
        self.browser = self.pw.chromium.launch(headless=False)
        self.page = self.browser.new_page()

    def snow_de(self, request_type = "Block/unblock"):
        self.page.goto(snow, wait_until="domcontentloaded")
        self.page.get_by_role("textbox", name="Enter your email, phone, or").fill("PIOTR.PIENKOWSKI@elancoah.com")
        self.page.keyboard.press("Enter")
        self.page.locator("#s2id_sp_formfield_sales_organization a").click()
        self.page.get_by_role("option", name="DE01").click()
        self.page.locator("#s2id_sp_formfield_type_of_request a").click()
        self.page.get_by_role("option", name=request_type).click()
        self.page.get_by_role("textbox", name="Customer Number").fill(self.CN)
        self.page.locator("#s2id_sp_formfield_request_priority a").click()
        self.page.get_by_role("option", name="Standard - 48 hours").click()
        self.page.locator("#s2id_sp_formfield_distribution_channel a").click()
        self.page.get_by_role("option", name="10").click()
        self.page.locator("#s2id_sp_formfield_multiple_requests a").click()
        self.page.get_by_role("option", name="No", exact=True).click()
        self.page.locator("#s2id_sp_formfield_account_group a").click()
        self.page.get_by_role("option", name="Sold-to").click()
        self.page.locator('#cmd_form_attached input[type="file"]').set_input_files(rf"{new_file}\{self.CN}.xlsm")

    def snow_ch(self,request_type = "Block/unblock"):
        self.page.goto(snow_ch, wait_until="domcontentloaded")
        self.page.get_by_role("textbox", name="Enter your email, phone, or").fill("PIOTR.PIENKOWSKI@elancoah.com")
        self.page.keyboard.press("Enter")
        self.page.locator("#s2id_sp_formfield_type_of_request a").click()
        self.page.get_by_role("option", name= request_type ).click()
        self.page.locator("#s2id_sp_formfield_account_group a").click()
        self.page.get_by_role("option", name="Z001 - Sold to").click()
        self.page.locator("#s2id_sp_formfield_region a").click()
        self.page.get_by_role("option", name="EMEA").click()
        self.page.locator("#s2id_sp_formfield_sales_organization_emea a").click()
        self.page.get_by_role("option", name="CH01").click()
        self.page.locator("#s2id_sp_formfield_request_priority a").click()
        self.page.get_by_role("option", name="Standard - 48 hours").click()
        self.page.locator("#s2id_sp_formfield_multiple_request a").click()
        self.page.get_by_role("option", name="No").click()
        ## tu powinien byc domestic
        self.page.locator('#cmd_form_attached input[type="file"]').set_input_files(rf"{new_file}\{self.CN}.xlsm")
        time.sleep(300)

class de(snow_ticket):

    def __init__(self, CN):
        super().__init__()
        self.CN = CN
        self.wb= load_workbook(template, keep_vba= True)
        self.ws = self.wb.worksheets[0]
        self.ws['A12'] = 'DE01'
        self.ws['B12'] = 'Block/Unblock/Deletion Flag'
        self.ws['C12'] = 'Sold-to'
        self.ws['E5'] = CN

    def zapisywanie (self):
        self.wb.save(rf'{new_file}\{self.CN}.xlsm')
        self.wb.close()

    def set_central_order_block(self):
        self.ws['E13'] = 'Set'
        self.ws['E14'] = '01 - Overall block'
        self.zapisywanie()
        self.snow_de()
        time.sleep(300)

    def remove_central_order_block(self):
        self.ws['E13'] = 'Reset (Remove)'
        self.ws['E14'] = '01 - Overall block'
        self.zapisywanie()
        self.snow_de()
        time.sleep(300)

    def remove_deletion_flag(self):
        self.ws['E12'] = 'Reset (Remove)'
        self.zapisywanie()
        self.snow_de("Deletion Flag")
        time.sleep(300)

    def remove_central_order_and_deletion_flag(self):
        self.ws['E12'] = 'Reset (Remove)'
        self.ws['E13'] = 'Reset (Remove)'
        self.ws['E14'] = '01 - Overall block'
        self.zapisywanie()
        self.snow_de("Deletion Flag")
        time.sleep(300)

class ch(de):

    def __init__(self, CN):
        super().__init__(CN)
        self.ws['A12'] = 'CH01'
        self.ws['B12'] = 'Block/Unblock/Deletion Flag'
        self.ws['C12'] = 'Sold-to'
        self.ws['E5'] = CN

    def set_central_order_block(self):
        self.ws['E13'] = 'Set'
        self.ws['E14'] = '01 - Overall block'
        self.zapisywanie()
        self.snow_ch()
        time.sleep(300)

    def remove_central_order_block(self):
        self.ws['E13'] = 'Reset (Remove)'
        self.ws['E14']= '01 - Overall block'
        self.zapisywanie()
        self.snow_ch()
        time.sleep(300)

    def remove_deletion_flag(self):
        self.ws['E12']= 'Reset (Remove)'
        self.zapisywanie()
        self.snow_ch("Deletion Flag")
        time.sleep(300)

    def remove_central_order_and_deletion_flag(self):
        self.ws['E12'] = 'Reset (Remove)'
        self.ws['E13'] = 'Reset (Remove)'
        self.ws['E14']= '01 - Overall block'
        self.zapisywanie()
        self.snow_ch("Deletion Flag")
        time.sleep(300)

de("2").remove_deletion_flag()




#  klasy de i ch
# set_central_order_block  - d
# remove_central_order_block
# remove_deletion_flag
# remove_central_order_and_deletion_flag