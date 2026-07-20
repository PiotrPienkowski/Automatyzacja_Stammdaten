from openpyxl import load_workbook
import warnings


warnings.filterwarnings(
"ignore",
message="Data Validation extension is not supported and will be removed"
)

class Licencje:
    def __init__(self, CN,):
        Template = rf'C:\Users\02703821\Elanco\CH - Bestellung Monitoring\Templatka do pythona GTS\(sold-to  change  DE01)   CMD_template4.1.4.xlsm'
        self.CN = CN
        self.wb = load_workbook(Template, keep_vba=True)
        self.ws= self.wb['Sheet1']

        self.__set_comon_fields()  # wywołanie metody klasy Licencje set Common)_fields

    def __set_comon_fields(self):
        self.ws['A12'] = 'DE01'
        self.ws['B12']= 'Change'
        self.ws['C12']= 'Sold-to'
        self.ws['E5'] = self.CN

    def close(self):
        sciezka = rf'C:\Users\02703821\OneDrive - Elanco\Desktop\robocze\{self.CN} create C33 licence.xlsm'
        self.wb.save(sciezka)
        self.wb.close()

    def C06(self):

        self.ws['E23'] = 'C06'
        self.ws['E59'] = 'Yes'
        self.ws['E60'] = 'C06 - Veterinary'
        self.ws['E61'] = 'NA'
        self.ws['E62'] = '30.12.9999'
        self.ws['E63'] = 'L01, L02, L03, L04, NONE'
        self.ws['E64'] = 'NA'
        self.ws['E65'] = '9,999,999'

        self.close()

    def C33(self):

        self.ws['E23'] = 'C33'
        self.ws['E59'] = 'Yes'
        self.ws['E60'] = 'C33 - Vet Samples'
        self.ws['E61'] = 'NA'
        self.ws['E62'] = '30.12.9999'
        self.ws['E63'] = 'NA'
        self.ws['E64'] = 'CA5536030GQZ1, CA5537030GQZ1, CA5538030GQZ1, CA5539030GQZ1'
        self.ws['E65'] = '2'

        self.close()

    def C34(self):

        self.ws['E23'] = 'C34'
        self.ws['E59'] = 'Yes'
        self.ws['E60'] = 'C34 - Registration for Complementary Feed for Farm Animals'
        self.ws['E61'] = 'NA'
        self.ws['E62'] = '30.12.9999'
        self.ws['E63'] = 'L10'
        self.ws['E64'] = 'NA'
        self.ws['E65'] = '9,999,999'

        self.close()


