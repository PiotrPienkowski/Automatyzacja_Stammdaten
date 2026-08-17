from IPython.core import page
from openpyxl import load_workbook
import warnings
import os
from playwright.sync_api import sync_playwright
import time

warnings.filterwarnings(
"ignore",
message="Data Validation extension is not supported and will be removed"
)
os.system("taskkill /F /IM excel.exe 1>nul 2>nul")

link_do_snow = rf'https://thespot.elanco.com/esc?id=sc_cat_item&sys_id=9d661f191b03d1105ca7eca3604bcb3a&sysparm_category=a20cb8eedb7c60905513c3af299619d0'
path = rf'C:\Users\02703821\Elanco\CH - Bestellung Monitoring\Templatka do pythona GTS\(sold-to  change  DE01)   CMD_template4.1.4.xlsm'

class licencja:

    def __init__(self,CN):
        self.CN = CN
        self.wb = load_workbook(path, keep_vba=True)
        self.ws = self.wb.worksheets[0]
        self.ws['A12'] = 'DE01'
        self.ws['B12'] = 'Change'
        self.ws['C12'] = 'Sold-to'
        self.ws['E5'] = CN

    def snow_ticket(self):
        self.pw = sync_playwright().start()
        self.browser = self.pw.chromium.launch(channel= "chromium", headless= False)
        self.tab = self.browser.new_page()
        self.tab.goto(link_do_snow)
        self.tab.locator("#s2id_sp_formfield_sales_organization a").click()
        self.tab.get_by_role("option", name="DE01").click()
        self.tab.locator("#s2id_sp_formfield_type_of_request a").click()
        self.tab.get_by_role("option", name="Change").click()
        self.tab.get_by_role("textbox", name="Customer Number").fill(self.CN)
        self.tab.locator("#s2id_sp_formfield_request_priority a").click()
        self.tab.get_by_role("option", name="Critical - 4 hours").click()
        self.tab.locator("#s2id_sp_formfield_distribution_channel a").click()
        self.tab.get_by_role("option", name="10-Domestic").click()
        self.tab.locator("#s2id_sp_formfield_multiple_requests a").click()
        self.tab.get_by_role("option", name="No", exact= True).click()
        self.tab.locator("#s2id_sp_formfield_multiple_requests a").click()
        self.tab.locator("#s2id_sp_formfield_account_group a").click()
        self.tab.get_by_role("option", name="Sold-to").click()
        self.tab.get_by_role("textbox", name="Additional information").fill(f'Hello Team, Please create C08 licence (See attached)')
        with self.tab.expect_file_chooser() as fc:
            self.tab.get_by_role("button", name = "Upload Attachment for CMD").click()
        file_chooser = fc.value
        file_chooser.set_files(self.sciezka)
        with self.tab.expect_file_chooser() as fca:
            self.tab.get_by_role("button", name = "Upload Attachment for VET").click()
        for i in os.listdir(r'C:\Users\02703821\OneDrive - Elanco\Desktop\robocze'):
            if i.endswith('.pdf'):
                license_chooser = fca.value
                license_chooser.set_files(os.path.join(r'C:\Users\02703821\OneDrive - Elanco\Desktop\robocze\\', i))
                break

    def C33(self):
        self.ws['E23'] =  'C33'
        self.ws['E59'] = 'Yes'
        self.ws['E60'] = 'C33 - Vet Samples'
        self.ws['E61'] = 'NA'
        self.ws['E62'] = '31.12.2026'
        self.ws['E63'] = 'NA'
        self.ws['E64'] = 'CA5536030GQZ1, CA5537030GQZ1, CA5538030GQZ1, CA5539030GQZ1'
        self.ws['E65'] = '2'
        self.sciezka = rf'C:\Users\02703821\OneDrive - Elanco\Desktop\robocze\{self.CN} create {self.ws["E23"].value} licence.xlsm'
        self.wb.save(self.sciezka)
        self.snow_ticket()
        time.sleep(300)

    def C06(self):
        self.ws['E23'] =  'C06'
        self.ws['E59'] = 'Yes'
        self.ws['E60'] = 'C06 - Veterinary'
        self.ws['E61'] = 'NA'
        self.ws['E62'] = '30.12.9999'
        self.ws['E63'] = 'L01, L02, L03, L04, NONE'
        self.ws['E64'] = 'NA'
        self.ws['E65'] = '9,999,999'
        self.sciezka = rf'C:\Users\02703821\OneDrive - Elanco\Desktop\robocze\{self.CN} create {self.ws["E23"].value} licence.xlsm'
        self.wb.save(self.sciezka)
        self.snow_ticket()
        time.sleep(300)

    def C34(self):
        self.ws['E23'] =  'C34'
        self.ws['E59'] = 'Yes'
        self.ws['E60'] = 'C34 - Registration for Complementary Feed for Farm Animals'
        self.ws['E61'] = 'NA'
        self.ws['E62'] =  '30.12.9999'
        self.ws['E63'] = 'L10'
        self.ws['E64'] = 'NA'
        self.ws['E65'] = '9,999,999'
        self.sciezka = rf'C:\Users\02703821\OneDrive - Elanco\Desktop\robocze\{self.CN} create {self.ws["E23"].value} licence.xlsm'
        self.wb.save(self.sciezka)
        self.snow_ticket()
        time.sleep(9999999)


licencja("123").C34()