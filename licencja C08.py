import win32com.client as win32
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from playwright.sync_api import sync_playwright
import os
import time

path = r'C:\Users\02703821\OneDrive - Elanco\Desktop\robocze'
link_do_snow = "https://thespot.elanco.com/esc?id=sc_cat_item&sys_id=9d661f191b03d1105ca7eca3604bcb3a&sysparm_category=a20cb8eedb7c60905513c3af299619d0"
tracker = r'C:\Users\02703821\Elanco\CH - Bestellung Monitoring\GTS Bestellungen (3).xlsx'

result = os.system("taskkill /F /IM excel.exe 1>nul 2>nul") ## >nul (czarna dziura nic nie wyswietla) - przekierowuje standardowe komunikaty do "kosza" (1),przekierowuje komunikaty błędów do kosza 2)
if result != 0:  #0 to jest polecenie wykonane poprawnie tzn. procesy zamkniete <>0 blad systemwy ale nie blad obslugiwany przez sxcept
    print("Nie znaleziono otwartego Excela")

def C08(CN, BTM):

    excel = win32.Dispatch('Excel.Application')
    excel.Visible = True
    wb = excel.Workbooks.Open(r'C:\Users\02703821\Elanco\CH - Bestellung Monitoring\CMD_template4.1.4.xlsm')
    ws = wb.Worksheets('Sheet1')
    ws.Range('A12').Value = 'DE01'
    ws.Range('B12').Value = 'Change'
    ws.Range('C12').Value = 'Sold-to'
    excel.Application.Run("CreatingHeader")
    excel.CalculateUntilAsyncQueriesDone()

    ws.Range('E5').Value = CN
    ws.Range('E23').Value = "C08"
    ws.Range('E59').Value = 'Yes'
    ws.Range('E60').Value = 'C08 - DEA Licence/Narcotic'
    ws.Range('E61').Value = BTM

    new_file = rf'C:\Users\02703821\OneDrive - Elanco\Desktop\robocze\{CN}_create C08 licence.xlsm'
    wb.CheckCompatibility = False
    wb.SaveAs(new_file)
    wb.Close(SaveChanges=False)

    wb1 = excel.Workbooks.Open(r'C:\Users\02703821\Elanco\CH - Bestellung Monitoring\BTM Template.xlsx')
    ws1 = wb1.Worksheets(1)

    wb1.RefreshAll()
    excel.CalculateUntilAsyncQueriesDone()
    data1 = ws1.UsedRange.Value
    df = pd.DataFrame(data1)

    # pierwszy wiersz -> nagłówki
    df.columns = df.iloc[0]

    # usuń pierwszy wiersz
    df = df.iloc[1:].reset_index(drop=True)
    df = df[df['Kundennummer'].astype(str).str.replace('.0', '', regex=False) == CN]
    df['KLIENT'] = '342'
    df['NR_BTM'] = BTM

    new_file1 = rf'C:\Users\02703821\OneDrive - Elanco\Desktop\robocze\pharmlog {CN}.xlsx'
    df.to_excel(new_file1, index=False)

    wb3 = load_workbook(new_file1)
    ws3 = wb3.active

    colour = PatternFill(fill_type='solid', fgColor='C0C0C0')

    for cell in ws3[1]:
        cell.fill = colour
        ws3.column_dimensions[cell.column_letter].width = 20

    for col in ['C', 'D', 'E', 'F']:
        ws3.column_dimensions[col].width = 40

    wb3.save(new_file1)
    wb3.close()
    wb1.Close(SaveChanges=False)

    # wb_tracker = excel.Workbooks.Open(tracker)
    # ws_tracker = wb_tracker.Sheets("Piotr- technical tab 2")
    # table = ws_tracker.ListObjects("Tabela3")
    # table.AutoFilter.ShowAllData()
    # table.Range.AutoFilter(Field=9, Criteria1= CN)


    outlook = win32.Dispatch('Outlook.Application')
    new_mail = outlook.CreateItem(0)
    new_mail.To = 'tl@pharmlog.de;btm@pharmlog.de'
    new_mail.CC = 'stammdaten@elancoah.com'
    new_mail.Subject = (rf'{CN} BTM')
    new_mail.Body = (f'Guten Tag,\n '
                     f'Anbei BTM \n\n\n\n '
                     f"Mit freundlichen Grüßen\n\n"
                     f"Piotr Pieńkowski\n"
                     f"Master Data Manager DACH\n\n"
                     f"Elanco Deutschland GmbH\n"
                     f"Rolf-Schwarz-Schütte-Platz 2 / CC A01 1.OG\n"
                     f"D-40789 Monheim am Rhein\n"
                     f"piotr.pienkowski@elancoah.com\n"
                     f"www.elanco.de\n\n"
                     f"AG Bad Homburg HRB 13307 | Geschäftsführung:\n"
                     f"Dr. Inga Drosse und Cindy Wichmann\n"
                     f"Sitz der Gesellschaft: Bad Homburg\n\n"
                     f"CONFIDENTIALITY NOTICE: This email message (including all attachments) "
                     f"is for the sole use of the intended recipient(s) and may contain confidential "
                     f"and privileged information. Any unauthorized review, use, disclosure, "
                     f"copying or distribution is strictly prohibited. If you are not the intended "
                     f"recipient, please contact the sender by reply email and destroy all copies "
                     f"of the original message.\n\n"
                     f"PRIVACY NOTICE: Your privacy is important to us. To find out more about "
                     f"the information that Elanco may collect, how we use it, how we protect it, "
                     f"and your rights and choices with respect to your Personal Data, go to "
                     f"privacy.elanco.com"
                     )

    new_mail.Attachments.Add(rf'C:\Users\02703821\OneDrive - Elanco\Desktop\robocze\pharmlog {CN}.xlsx')

    for file in os.listdir(path):
        if file.lower().endswith('.pdf'):
            new_mail.Attachments.Add(os.path.join(path, file))

    new_mail.Display()

    start = sync_playwright().start()
    browser = start.chromium.launch(headless= False)
    page = browser.new_page()
    page.goto(link_do_snow,wait_until="domcontentloaded")
    page.get_by_role("textbox", name="Enter your email, phone, or").fill("PIOTR.PIENKOWSKI@elancoah.com")
    page.keyboard.press("Enter")
    page.locator("#s2id_sp_formfield_sales_organization a").click()
    page.get_by_role("option", name="DE01").click()
    page.locator("#s2id_sp_formfield_type_of_request a").click()
    page.get_by_role("option", name="Change").click()
    page.get_by_role("textbox", name="Customer Number").fill(CN)
    page.locator("#s2id_sp_formfield_request_priority a").click()
    page.get_by_role("option", name="Critical - 4 hours").click()
    page.locator("#s2id_sp_formfield_distribution_channel a").click()
    page.get_by_role("option", name="10-Domestic").click()
    page.locator("#s2id_sp_formfield_multiple_requests a").click()
    page.get_by_role("option", name="No", exact=True).click()
    page.locator("#s2id_sp_formfield_account_group a").click()
    page.get_by_role("option", name="Sold-to").click()
    page.get_by_role("button", name = "Upload Attachment for VET").click()
    page.get_by_role("textbox", name = "Additional information").fill(f'Hello Team, Please create C08 licence (See attached)')

    with page.expect_file_chooser() as fc:
        page.get_by_role(
            "button",
            name = "Upload Attachment for CMD").click()
    fc.value.set_files(new_file)

    with page.expect_file_chooser()as cf1:
        page.get_by_role("button", name="Upload Attachment for VET").click()

    for i in os.listdir(r'C:\Users\02703821\OneDrive - Elanco\Desktop\robocze'):
        if os.path.join(r'C:\Users\02703821\OneDrive - Elanco\Desktop\robocze', i).endswith(('.pdf','.jpg', '.png')):
            cf1.value.set_files(os.path.join(r'C:\Users\02703821\OneDrive - Elanco\Desktop\robocze', i))
            break

    time.sleep(99999)

C08('50884388','4667675')