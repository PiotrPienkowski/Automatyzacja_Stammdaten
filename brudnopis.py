#
# from playwright.sync_api import sync_playwright
# import time
#
# with sync_playwright() as p:
#     browser = p.chromium.launch(headless=False)
#     page = browser.new_page()
#     page.goto("https://login.veevanetwork.com/auth/login?retURL=https%3A%2F%2Felanco.veevanetwork.com/ui/",wait_until="domcontentloaded")
#     page.pause()

# from playwright.sync_api import sync_playwright
#
# with sync_playwright() as p:
#     context = p.chromium.launch_persistent_context(
#     user_data_dir="veeva_profile",
#     headless=False
#     )
#     page = context.new_page()
#     page.goto(
#     "https://elanco.veevanetwork.com/ui/",
#     wait_until="networkidle"
#     )
#     page.pause()

# from playwright.sync_api import sync_playwright
# import time
#
# playwright = sync_playwright().start()
# browser = playwright.chromium.launch(headless=False)
# page = browser.new_page()
# page.goto("https://login.veevanetwork.com/auth/login?retURL=https%3A%2F%2Felanco.veevanetwork.com/ui/")
# page.wait_for_load_state("networkidle")
# page.get_by_text("Log in with").click()
# time.sleep(300)

# from os import path

import pandas as pd
import playwright.async_api
import win32com
# import os
#
# prath = rf'C:\Users\02703821\OneDrive - Elanco\Desktop\robocze'
#
# for i in os.listdir(prath):
#     if os.path.join(prath,i).endswith('.pdf'):
#         print(i)
#     else:
#         print("png")

# import xlsxwriter as xl
# import pandas as pd
#
# path = rf'C:\Users\02703821\OneDrive - Elanco\Desktop\robocze\pharmlog 50022351.xlsx'
#
# df = pd.read_excel(path,sheet_name='Sheet1')
# numer= df[df["Kundennummer"] == 50022351]
# print(numer)
#

# import win32com.client as win32com
# import os
# import time
# import pandas as pd
#
# os.system('taskkill /f /im excel.exe')
#
# tracker = r'C:\Users\02703821\Elanco\CH - Bestellung Monitoring\GTS Bestellungen (3).xlsx'
#
# excel = win32com.Dispatch("Excel.Application")
# excel.visible= True
# wb_tracker = excel.Workbooks.Open(tracker)
# ws_tracker = wb_tracker.Sheets("Piotr- technical tab 2")
# table = ws_tracker.ListObjects("Tabela3")
# table.AutoFilter.ShowAllData()
# table.Range.AutoFilter(Field= 9, Criteria1 = "50000779")
# time.sleep(300)

# veeva = r"https://login.veevanetwork.com/auth/login?retURL=https%3A%2F%2Felanco.veevanetwork.com/ui/"
#
# playwright = sync_playwright().start()
# browser = playwright.chromium.launch(channel= "chromium", headless= False)
# page = browser.new_page()
# page.goto(veeva)
# page.get_by_role("button", name="Log in with Microsoft").click()
# time.sleep(300)


# from playwright.sync_api import sync_playwright
#
# playwright = sync_playwright().start()
# browser = playwright.chromium.launch(headless=False)
# tab = browser.new_page()
# tab.goto("https://login.veevanetwork.com/auth/login?retURL=https%3A%2F%2Felanco.veevanetwork.com/ui/")
# tab.get_by_role("button", name = "Log in with").click()


# from playwright.sync_api import sync_playwright
# import time
#
#
# playwright = sync_playwright().start()
# browser = playwright.chromium.launch(headless=False)
# page = browser.new_page()
# page.goto("https://elancoconnect.lightning.force.com/lightning/page/home")
# time.sleep(300)


from openpyxl import load_workbook
import time
import os

# os.system("taskkill /F /IM excel.exe 1>nul 2>nul")
#
# formatka = r'C:\Users\02703821\Elanco\CH - Bestellung Monitoring\CMD_template4.1.4.xlsm'
# nowy_plik = r'C:\Users\02703821\OneDrive - Elanco\Desktop\robocze\test.xlsm'
#
# def funkcja(CN):
#
#     excel = load_workbook(formatka, keep_vba=True)
#     ws = excel.worksheets[0]
#     ws['A12'] = 'CH01'
#     ws['B12'] = 'Block/Unblock/Deletion Flag'
#     ws['C12'] = 'Sold-to'
#     ws['E5'] = CN
#     ws['E23']= "test"
#
#     excel.save(nowy_plik)
#     excel.close()
#
# funkcja('moninski')


# df = pd.read_excel(r'C:\Users\02703821\Elanco\CH - Salesforce Source File\EAHSAP DE & CH mit Details (Piotr).xlsx')
#
# prompt = f"""
# Przeanalizuj dane klientów:
# {df.head(20).to_string()}
# 10
# """
#
# response = ollama.chat(
# model="llama3",
# messages=[
# {"role": "user", "content": prompt}
# ]
# )
# print(response["message"]["content"])

# from playwright.sync_api import sync_playwright
# import time
#
# with sync_playwright() as p:
#     browser = p.chromium.launch(headless=False)
#     page = browser.new_page()
#     page.goto("https://thespot.elanco.com/esc?id=sc_cat_item&sys_id=9d661f191b03d1105ca7eca3604bcb3a&sysparm_category=a20cb8eedb7c60905513c3af299619d0")
#     page.pause()
#

# from playwright.sync_api import sync_playwright
#
#
# pl = sync_playwright().start()
# chrome = pl.chromium.launch(headless = False)
# page = chrome.new_page()
# page.goto("https://thespot.elanco.com/esc?id=sc_cat_item&sys_id=9d661f191b03d1105ca7eca3604bcb3a&sysparm_category=a20cb8eedb7c60905513c3af299619d0",wait_until="domcontentloaded")
# page.get_by_role("textbox", name="Enter your email, phone, or").fill("PIOTR.PIENKOWSKI@elancoah.com")
# page.keyboard.press("Enter")
# input("Naciśnij Enter po zalogowaniu...")

from playwright.sync_api import sync_playwright

# with sync_playwright() as p:
#
#     context = p.chromium.launch_persistent_context(
#          user_data_dir=r"C:\PlaywrightProfile",
#          channel="chrome",
#          headless=False
# )
#     page = context.new_page()
#     page.goto("https://thespot.elanco.com/esc?id=sc_cat_item&sys_id=9d661f191b03d1105ca7eca3604bcb3a&sysparm_category=a20cb8eedb7c60905513c3af299619d0")
#     input("Enter...")

# from playwright.sync_api import sync_playwright
#
# pl = sync_playwright().start()
# context = pl.chromium.launch_persistent_context(user_data_dir=r"C:\PlaywrightProfile",channel="chrome",headless=False)
# page = context.new_page()
# page.goto("https://thespot.elanco.com/esc?id=sc_cat_item&sys_id=9d661f191b03d1105ca7eca3604bcb3a&sysparm_category=a20cb8eedb7c60905513c3af299619d0")
# input("Enter...")


# from playwright.sync_api import sync_playwright
#
# with sync_playwright() as p:
#     context = p.chromium.launch_persistent_context(
#     user_data_dir=r"C:\Users\02703821\playwright_profile",
#     headless=False
#     )
#
#     page = context.new_page()
#
#     page.goto("https://thespot.elanco.com")
#
#     input("Zaloguj się ręcznie i naciśnij Enter...")
#
#     context.close()

# import time
# from pywinauto import Desktop
# import win32com.client as win32
#
# path = r'C:\Users\02703821\Elanco\CH - Bestellung Monitoring\BTM Template.xlsx'
#
# excel = win32.Dispatch("Excel.Application")
# excel.Visible = True
# wb = excel.Workbooks.Open(path)
# ws = wb.Worksheets(1)
# wb.RefreshAll()
#
# time.sleep(5)
#
# dlg = Desktop(backend="uia").window(title_re=".*Pick an account.*")
# dlg.wait("visible", timeout=30)
#
# print("okno znalezione")

# locator("div").nth(5)
# get_by_text("Log in with")

from playwright.sync_api import sync_playwright
import time

CN = "0050727431"

with sync_playwright() as p:
    context = p.chromium.launch_persistent_context(user_data_dir="veeva_profile",headless=False)
    page = context.new_page()
    page.goto("https://elanco.veevanetwork.com/ui/",wait_until="networkidle")
    # page.get_by_role("textbox", name ="User Name").fill("PIOTR.PIENKOWSKI@elancoah.com")
    # page.get_by_role("textbox", name ="Password").fill("4117Rkxe!!!!")
    # page.locator("div").nth(5).click()
    # page.get_by_role("textbox",name= "Enter your email, phone, or").fill("PIOTR.PIENKOWSKI@elancoah.com")
    # page.get_by_role("button", name = "Next").click()
    search = page.locator(".input").first
    search.click()
    search.fill(CN)
    search.press("Enter")
    time.sleep(9999)

