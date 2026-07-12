import win32com.client as win32
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

excel = win32.DispatchEx('Excel.Application')
excel.Visible = False
wb = excel.Workbooks.Open(r'C:\Users\02703821\Elanco\CH - Bestellung Monitoring\CMD_template4.1.4.xlsm')
ws = wb.Worksheets('Sheet1')
ws.Range('A12').Value = 'DE01'
ws.Range('B12').Value = 'Change'
ws.Range('C12').Value = 'Sold-to'
excel.Application.Run("CreatingHeader")

excel.CalculateUntilAsyncQueriesDone()

def C08(CN,BTM):

# runowanie makra vba

    ws.Range('E5').Value = CN
    ws.Range('E23').Value = "C08"
    ws.Range('E59').Value = 'Yes'
    ws.Range('E60').Value = 'C08 - DEA Licence/Narcotic'
    ws.Range('E61').Value = BTM

    new_file = rf'C:\Users\02703821\OneDrive - Elanco\Desktop\robocze\{CN}_create C08 licence.xlsm'
    wb.SaveAs(new_file)
    wb.Close(SaveChanges=False)


    wb1 = excel.Workbooks.Open(r'C:\Users\02703821\Elanco\CH - Bestellung Monitoring\BTM Template.xlsx')
    ws1 = wb1.Worksheets(1)
    wb1.RefreshAll()
    excel.CalculateUntilAsyncQueriesDone()
    data1 = ws1.UsedRange.Value
    df = pd.DataFrame(data1)

    # pierwszy wiersz -> nagłówki
    df.columns = df.iloc[0]

    # usuń pierwszy wiersz
    df = df.iloc[1:].reset_index(drop=True)
    df = df[df['Kundennummer'].astype(str).str.replace('.0', '', regex=False) == CN]
    df['KLIENT'] = '342'
    df['NR_BTM'] = BTM


    new_file1 = rf'C:\Users\02703821\OneDrive - Elanco\Desktop\robocze\pharmlog {CN}.xlsx'
    df.to_excel(new_file1, index=False)

    wb3 =load_workbook(new_file1)
    ws3 = wb3.active

    colour = PatternFill(fill_type='solid', fgColor='C0C0C0')

    for cell in ws3[1]:
        cell.fill = colour
        ws3.column_dimensions[cell.column_letter].width = 20

    for col in ['C', 'D', 'E', 'F']:
        ws3.column_dimensions[col].width = 40

    wb3.save(new_file1)
    wb3.close()

    wb1.Close(SaveChanges=False)




    outlook = win32.Dispatch('Outlook.Application')
    new_mail = outlook.CreateItem(0)
    new_mail.To = 'tl@pharmlog.de;btm@pharmlog.de'
    new_mail.CC = 'stammdaten@elancoah.com'
    new_mail.Subject = (rf'{CN} BTM')
    new_mail.Body = (f'Guten Tag,\n '
                     f'Anbei BTM \n\n\n\n '
                     f"Mit freundlichen Grüßen\n\n"
                     f"Piotr Pieńkowski\n"
                     f"Master Data Manager DACH\n\n"
                     f"Elanco Deutschland GmbH\n"
                     f"Rolf-Schwarz-Schütte-Platz 2 / CC A01 1.OG\n"
                     f"D-40789 Monheim am Rhein\n"
                     f"piotr.pienkowski@elancoah.com\n"
                     f"www.elanco.de\n\n"
                     f"AG Bad Homburg HRB 13307 | Geschäftsführung:\n"
                     f"Dr. Inga Drosse und Cindy Wichmann\n"
                     f"Sitz der Gesellschaft: Bad Homburg\n\n"
                     f"CONFIDENTIALITY NOTICE: This email message (including all attachments) "
                     f"is for the sole use of the intended recipient(s) and may contain confidential "
                     f"and privileged information. Any unauthorized review, use, disclosure, "
                     f"copying or distribution is strictly prohibited. If you are not the intended "
                     f"recipient, please contact the sender by reply email and destroy all copies "
                     f"of the original message.\n\n"
                     f"PRIVACY NOTICE: Your privacy is important to us. To find out more about "
                     f"the information that Elanco may collect, how we use it, how we protect it, "
                     f"and your rights and choices with respect to your Personal Data, go to "
                     f"privacy.elanco.com"
                     )
    path = r'C:\Users\02703821\OneDrive - Elanco\Desktop\robocze'
    new_mail.Attachments.Add(rf'C:\Users\02703821\OneDrive - Elanco\Desktop\robocze\pharmlog {CN}.xlsx')

    for file in os.listdir(path):
        if file.lower().endswith('.pdf'):
            new_mail.Attachments.Add(os.path.join(path, file))

    new_mail.Display()

def C06(CN):


    ws.Range('E5').Value = CN
    ws.Range('E23').Value = "C06"
    ws.Range('E59').Value = 'Yes'
    ws.Range('E60').Value = 'C06 - Veterinary'

    path = rf'C:\Users\02703821\OneDrive - Elanco\Desktop\robocze\{CN} create C06 licence.xlsm'
    wb.SaveAs(path)
    wb.Close(SaveChanges=False)


def C34(CN):

    ws.Range('E5').Value = CN
    ws.Range('E23').Value = "C34"
    ws.Range('E59').Value = 'Yes'
    ws.Range('E60').Value = 'C34 - Registration for Complementary Feed for Farm Animals'

    path = rf'C:\Users\02703821\OneDrive - Elanco\Desktop\robocze\{CN} create C34 licence.xlsm'
    wb.SaveAs(path)
    wb.Close(SaveChanges=False)

def C33(CN):


    ws.Range('E5').Value = CN
    ws.Range('E23').Value = "C33"
    ws.Range('E59').Value = 'Yes'
    ws.Range('E60').Value = 'C33 - Vet Samples'

    path = rf'C:\Users\02703821\OneDrive - Elanco\Desktop\robocze\{CN} create C33 licence.xlsm'
    wb.SaveAs(path)
    wb.Close(SaveChanges=False)


#######################



# haslo = input('jaka chcesz zalozyc licencje: ').strip().upper()
#
# if haslo == 'C08':
#     CN = input("podaj numer klienta: ")
#     BTM = input("podaj BTM: ")
#     C08(CN,BTM)
#
# elif haslo == 'C06':
#     CN = input("podaj numer klienta: ")
#     C06(CN)
#
# elif haslo == 'C33':
#     CN = input("podaj numer klienta: ")
#     C33(CN)
#
# elif haslo == 'C34':
#     CN = input("podaj numer klienta: ")
#     C34(CN)
#
# else:
#     print('podales zly numer licencji')
# excel.Quit()
#
